"""
下載 00981A ETF 最新持股 Excel，確認日期為當天後，
產生含收盤價的 00981A_YYYYMMDD.txt 至 00981A_Data 資料夾，
並更新 dates.json 供前端網頁動態載入日期清單。
"""

import os
import json
import requests
import warnings
from io import BytesIO
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ===== 設定 =====
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(BASE_DIR, "00981A_Data")
WORKDAY_FILE = os.path.join(BASE_DIR, "workday.txt")
LOG_FILE     = os.path.join(BASE_DIR, "download_log.txt")
DATES_JSON   = os.path.join(BASE_DIR, "dates.json")
URL          = "https://www.ezmoney.com.tw/ETF/Fund/AssetExcelNPOI?fundCode=49YTW"
PREFIX       = "00981A"

FINMIND_TOKEN = (
    "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9"
    ".eyJkYXRlIjoiMjAyNS0xMC0xOCAyMDo0Nzo0NSIsInVzZXJfaWQiOiJzZW5pb3IiLCJpcCI6IjExNC40NC4yNTEuNTkiLCJleHAiOjE3NjEzOTY0NjV9"
    ".QlkvCtdhOo-vgXgV6zF3RuA0VrjVyLTEh63aepHOsQ4"
)


# ── LOG ───────────────────────────────────────────────────────────────────
def log(msg: str):
    """同時輸出到 stdout 與 download_log.txt（附時間戳記）。"""
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ── 營業日 ────────────────────────────────────────────────────────────────
def get_workdays() -> set:
    """讀取 workday.txt，回傳 date 集合。"""
    workdays = set()
    if not os.path.exists(WORKDAY_FILE):
        log(f"⚠️ 找不到 {WORKDAY_FILE}")
        return workdays
    with open(WORKDAY_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    workdays.add(datetime.strptime(line, "%Y/%m/%d").date())
                except ValueError:
                    pass
    return workdays


# ── 收盤價 ────────────────────────────────────────────────────────────────
def get_close_price(stock_id: str, date_str: str) -> float:
    """以 FinMind API 取得指定日期收盤價。date_str 格式：YYYY-MM-DD"""
    url = "https://api.finmindtrade.com/api/v4/data"
    params = {
        "dataset":    "TaiwanStockPrice",
        "data_id":    stock_id,
        "start_date": date_str,
        "end_date":   date_str,
        #"token":      FINMIND_TOKEN,
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json().get("data", [])
        if data:
            return float(data[0]["close"])
        log(f"  ⚠️ {stock_id} 查無收盤價")
        return 0.0
    except Exception as e:
        log(f"  ❌ {stock_id} 取得收盤價失敗: {e}")
        return 0.0


# ── 下載 Excel ────────────────────────────────────────────────────────────
def download_excel() -> tuple[BytesIO | None, date | None]:
    """下載 Excel，解析 A1 取得資料日期（民國 → 西元）。"""
    log("📥 正在下載 Excel...")
    try:
        resp = requests.get(URL, stream=True, verify=False, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        log(f"❌ 下載失敗: {e}")
        return None, None

    raw = BytesIO(resp.content)
    try:
        wb = load_workbook(raw, data_only=True)
    except Exception as e:
        log(f"❌ 開啟 Excel 失敗: {e}")
        return None, None

    ws = wb.active
    a1 = str(ws["A1"].value or "").strip()
    # A1 格式範例：「資料日期：114/4/2」
    if "：" not in a1:
        log(f"❌ A1 格式不符: {a1!r}")
        return None, None

    date_part = a1.split("：", 1)[1].strip()
    try:
        roc_year, month, day = date_part.split("/")
        file_date = date(int(roc_year) + 1911, int(month), int(day))
    except Exception as e:
        log(f"❌ 日期解析失敗 ({date_part!r}): {e}")
        return None, None

    log(f"📅 Excel 資料日期: {file_date}")
    return raw, file_date


# ── 解析持股 ──────────────────────────────────────────────────────────────
def parse_holdings(raw: BytesIO) -> pd.DataFrame:
    """從 Excel BytesIO 解析持股資料（從第 21 行開始）。"""
    wb = load_workbook(raw, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=21, min_col=1, max_col=4, values_only=True):
        code, name, shares, weight = row
        if not code:
            break
        code_str   = str(code).strip().rstrip("*")
        name_str   = str(name).strip() if name else ""
        shares_str = str(shares).replace(",", "").strip() if shares else "0"
        weight_str = str(weight).strip() if weight else "0%"

        if weight_str and not weight_str.endswith("%"):
            try:
                weight_str = f"{float(weight_str)*100:.2f}%"
            except ValueError:
                pass

        rows.append({
            "股票代號": code_str,
            "股票名稱": name_str,
            "持股權重": weight_str,
            "股數":     shares_str,
        })

    return pd.DataFrame(rows)


# ── 儲存 TXT ──────────────────────────────────────────────────────────────
def save_txt(df: pd.DataFrame, file_date: date) -> str:
    """加入收盤價並儲存至 00981A_Data/00981A_YYYYMMDD.txt。"""
    os.makedirs(DATA_DIR, exist_ok=True)
    date_str     = file_date.strftime("%Y%m%d")
    date_api_str = file_date.strftime("%Y-%m-%d")
    save_path    = os.path.join(DATA_DIR, f"{PREFIX}_{date_str}.txt")

    if os.path.exists(save_path):
        log(f"ℹ️ 檔案已存在，將覆寫: {save_path}")

    log("📊 正在查詢各股收盤價...")
    prices = []
    total  = len(df)
    for i, row in df.iterrows():
        sid   = row["股票代號"]
        price = get_close_price(sid, date_api_str)
        prices.append(price)
        log(f"  [{i+1:>2}/{total}] {sid} {row['股票名稱']}: {price}")

    df = df.copy()
    df["收盤價"] = prices

    header = "股票代號,股票名稱,持股權重,股數,收盤價"
    lines  = [header]
    for _, row in df.iterrows():
        line = f"{row['股票代號']},{row['股票名稱']},{row['持股權重']},{row['股數']},{row['收盤價']}"
        lines.append(line)

    with open(save_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return save_path


# ── 更新 dates.json ───────────────────────────────────────────────────────
def update_dates_json():
    """掃描 00981A_Data 資料夾，產生日期清單寫入 dates.json。"""
    if not os.path.exists(DATA_DIR):
        return
    dates = sorted([
        f[7:15]                          # 取 YYYYMMDD 部分
        for f in os.listdir(DATA_DIR)
        if f.startswith(f"{PREFIX}_") and f.endswith(".txt") and len(f) == 19
    ])
    with open(DATES_JSON, "w", encoding="utf-8") as f:
        json.dump(dates, f)
    log(f"📋 dates.json 已更新，共 {len(dates)} 個交易日")


# ── 主程式 ────────────────────────────────────────────────────────────────
def main():
    log("=" * 50)
    today = date.today()
    log(f"📆 今天日期: {today}")

    # 確認今天是營業日
    workdays = get_workdays()
    if today not in workdays:
        log(f"⛔ {today} 不是營業日，結束。")
        #return

    # 下載並取得資料日期
    raw, file_date = download_excel()
    if raw is None:
        return

    # 比對資料日期是否為當天
    if file_date != today:
        log(f"⛔ Excel 資料日期 ({file_date}) 不是今天 ({today})，結束。")
        return

    log("✅ 日期符合，開始解析持股資料...")
    df = parse_holdings(raw)
    if df.empty:
        log("❌ 解析結果為空，請確認 Excel 格式。")
        return

    log(f"   共 {len(df)} 筆持股")

    # 查詢收盤價並儲存
    save_path = save_txt(df, file_date)
    log(f"✅ 完成！檔案已儲存: {save_path}")

    # 更新日期清單
    update_dates_json()


if __name__ == "__main__":
    main()
